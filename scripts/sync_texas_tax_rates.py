#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import tempfile
import urllib.request
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

YEAR = 2025
BASE = "https://comptroller.texas.gov/taxes/property-tax/docs"
FILES = {
    "county": f"{BASE}/{YEAR}-county-rates-levies.xlsx",
    "city": f"{BASE}/{YEAR}-city-rates-levies.xlsx",
    "school": f"{BASE}/{YEAR}-school-district-rates-levies.xlsx",
    "special": f"{BASE}/{YEAR}-special-district-rates-levies.xlsx",
}
OUTPUT = Path("src/data/counties.ts")
HOMESTEAD_EXEMPTION = 140_000

COUNTIES = [
"Anderson","Andrews","Angelina","Aransas","Archer","Armstrong","Atascosa","Austin","Bailey","Bandera","Bastrop","Baylor","Bee","Bell","Bexar","Blanco","Borden","Bosque","Bowie","Brazoria","Brazos","Brewster","Briscoe","Brooks","Brown","Burleson","Burnet","Caldwell","Calhoun","Callahan","Cameron","Camp","Carson","Cass","Castro","Chambers","Cherokee","Childress","Clay","Cochran","Coke","Coleman","Collin","Collingsworth","Colorado","Comal","Comanche","Concho","Cooke","Coryell","Cottle","Crane","Crockett","Crosby","Culberson","Dallam","Dallas","Dawson","Deaf Smith","Delta","Denton","DeWitt","Dickens","Dimmit","Donley","Duval","Eastland","Ector","Edwards","Ellis","El Paso","Erath","Falls","Fannin","Fayette","Fisher","Floyd","Foard","Fort Bend","Franklin","Freestone","Frio","Gaines","Galveston","Garza","Gillespie","Glasscock","Goliad","Gonzales","Gray","Grayson","Gregg","Grimes","Guadalupe","Hale","Hall","Hamilton","Hansford","Hardeman","Hardin","Harris","Harrison","Hartley","Haskell","Hays","Hemphill","Henderson","Hidalgo","Hill","Hockley","Hood","Hopkins","Houston","Howard","Hudspeth","Hunt","Hutchinson","Irion","Jack","Jackson","Jasper","Jeff Davis","Jefferson","Jim Hogg","Jim Wells","Johnson","Jones","Karnes","Kaufman","Kendall","Kenedy","Kent","Kerr","Kimble","King","Kinney","Kleberg","Knox","La Salle","Lamar","Lamb","Lampasas","Lavaca","Lee","Leon","Liberty","Limestone","Lipscomb","Live Oak","Llano","Loving","Lubbock","Lynn","Madison","Marion","Martin","Mason","Matagorda","Maverick","McCulloch","McLennan","McMullen","Medina","Menard","Midland","Milam","Mills","Mitchell","Montague","Montgomery","Moore","Morris","Motley","Nacogdoches","Navarro","Newton","Nolan","Nueces","Ochiltree","Oldham","Orange","Palo Pinto","Panola","Parker","Parmer","Pecos","Polk","Potter","Presidio","Rains","Randall","Reagan","Real","Red River","Reeves","Refugio","Roberts","Robertson","Rockwall","Runnels","Rusk","Sabine","San Augustine","San Jacinto","San Patricio","San Saba","Schleicher","Scurry","Shackelford","Shelby","Sherman","Smith","Somervell","Starr","Stephens","Sterling","Stonewall","Sutton","Swisher","Tarrant","Taylor","Terrell","Terry","Throckmorton","Titus","Tom Green","Travis","Trinity","Tyler","Upshur","Upton","Uvalde","Val Verde","Van Zandt","Victoria","Walker","Waller","Ward","Washington","Webb","Wharton","Wheeler","Wichita","Wilbarger","Willacy","Williamson","Wilson","Winkler","Wise","Wood","Yoakum","Young","Zapata","Zavala"
]

def normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()

def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")

def numeric(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group(0)) if match else 0.0

def download(url: str, target: Path) -> None:
    request = urllib.request.Request(url, headers={"User-Agent": "KeepTXRed tax-rate refresh"})
    with urllib.request.urlopen(request, timeout=90) as response:
        target.write_bytes(response.read())

def rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    raw = list(sheet.iter_rows(values_only=True))
    header_index = next((index for index, row in enumerate(raw[:20]) if sum(bool(cell) for cell in row) >= 3), 0)
    headers = [normalize(cell) or f"column {index}" for index, cell in enumerate(raw[header_index])]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in raw[header_index + 1 :]
        if any(value not in (None, "") for value in row)
    ]

def pick(record: dict[str, Any], *needles: str) -> Any:
    for needle in needles:
        for key, value in record.items():
            if needle in key:
                return value
    return None

def county_name(record: dict[str, Any]) -> str | None:
    raw = str(pick(record, "county name", "county") or "").replace(" County", "").strip()
    normalized = normalize(raw)
    return next((name for name in COUNTIES if normalize(name) == normalized), None)

def district_kind(name: str) -> str:
    upper = name.upper()
    if "MUD" in upper or "MUNICIPAL UTILITY" in upper or "WATER" in upper:
        return "MUD"
    if "PID" in upper or "PUBLIC IMPROVEMENT" in upper:
        return "PID"
    if "HOSPITAL" in upper:
        return "hospital"
    if "COLLEGE" in upper:
        return "college"
    if "EMERGENCY" in upper or " ESD" in upper:
        return "ESD"
    return "other"

def main() -> None:
    data = {
        name: {"countyRate": 0.0, "cities": [], "schools": [], "specials": []}
        for name in COUNTIES
    }

    with tempfile.TemporaryDirectory() as temp:
        temp_dir = Path(temp)
        parsed: dict[str, list[dict[str, Any]]] = {}
        for category, url in FILES.items():
            path = temp_dir / f"{category}.xlsx"
            download(url, path)
            parsed[category] = rows(path)

    for record in parsed["county"]:
        county = county_name(record)
        if county:
            data[county]["countyRate"] = numeric(pick(record, "total tax rate", "tax rate", "rate"))

    for record in parsed["city"]:
        county = county_name(record)
        name = str(pick(record, "city name", "taxing unit name", "name") or "").strip(" *")
        rate = numeric(pick(record, "total tax rate", "tax rate", "rate"))
        if county and name:
            data[county]["cities"].append({"name": name, "rate": rate})

    for record in parsed["school"]:
        county = county_name(record)
        name = str(pick(record, "school district name", "district name", "taxing unit name", "name") or "").strip(" *")
        rate = numeric(pick(record, "total tax rate", "tax rate", "rate"))
        if county and name:
            data[county]["schools"].append({"name": name, "rate": rate})

    for record in parsed["special"]:
        county = county_name(record)
        name = str(pick(record, "special district name", "district name", "taxing unit name", "name") or "").strip(" *")
        rate = numeric(pick(record, "total tax rate", "tax rate", "rate"))
        if county and name and rate > 0:
            data[county]["specials"].append({"name": name, "rate": rate, "kind": district_kind(name)})

    entries = []
    for name in COUNTIES:
        item = data[name]
        city_average = sum(city["rate"] for city in item["cities"]) / len(item["cities"]) if item["cities"] else 0.0
        special_average = sum(unit["rate"] for unit in item["specials"]) / len(item["specials"]) if item["specials"] else 0.0
        schools = item["schools"] or [{"name": "Enter exact ISD rate", "rate": 0.0}]
        entries.append({
            "slug": slug(name),
            "name": f"{name} County",
            "region": "Texas",
            "countyRate": round(item["countyRate"], 6),
            "cityAvgRate": round(city_average, 6),
            "specialDistrictRate": round(special_average, 6),
            "homesteadExemption": HOMESTEAD_EXEMPTION,
            "schoolDistricts": sorted(schools, key=lambda value: value["name"]),
            "specialDistricts": sorted(item["specials"], key=lambda value: value["name"]),
            "taxYear": YEAR,
            "dataSource": "Texas Comptroller PTAD",
        })

    types = (
        "// Generated by scripts/sync_texas_tax_rates.py. Do not edit manually.\n"
        "export type SchoolDistrict={name:string;rate:number};\n"
        "export type SpecialDistrict={name:string;rate:number;kind:\"MUD\"|\"PID\"|\"hospital\"|\"college\"|\"ESD\"|\"other\"};\n"
        "export type County={slug:string;name:string;region:string;countyRate:number;cityAvgRate:number;specialDistrictRate:number;homesteadExemption:number;schoolDistricts:SchoolDistrict[];specialDistricts:SpecialDistrict[];taxYear:number;dataSource:string};\n"
    )
    metadata = (
        f'export const TAX_RATE_DATASET={{taxYear:{YEAR},lastUpdated:"{date.today().isoformat()}",'
        f'sourceUrl:"https://comptroller.texas.gov/taxes/property-tax/rates/",countyCount:254}} as const;\n'
    )
    payload = json.dumps(entries, separators=(",", ":"), ensure_ascii=False)
    OUTPUT.write_text(
        types + metadata + f"export const COUNTIES:County[]={payload};\n"
        + "export function getCountyBySlug(value:string){return COUNTIES.find((county)=>county.slug===value)}\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(entries)} counties to {OUTPUT}")

if __name__ == "__main__":
    main()
